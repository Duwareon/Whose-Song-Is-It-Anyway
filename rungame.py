#!/usr/bin/env python3
print ("Starting Game...")

from uriformat import *
import openpyxl as xl
from os import system
from random import shuffle

# load the spreadsheet
wb = xl.load_workbook(filename = "songlist.xlsx")
ws = wb["Form Responses 1"]

playlist = []

for row in range(2, ws.max_row+1):
    for column in "BCDEFG":
        cell_name = "{}{}".format(column, row)
        if (cell_name[0] == "B"):
            ownername = ws[cell_name].value
        else:
            x = ws[cell_name].value
            if(type(ws[cell_name].value) == int):
               x = str(ws[cell_name].value) + 'a'
            playlist.append(Songlink(URLtoURI(x), ownername))

# remove duplicates and shuffle
for x in playlist:
    for y in playlist:
        #x.printData()
        #y.printData()
        if (x is y):
            pass
        elif(x.link == y.link):
            tempowner = y.owner
            playlist.remove(y)
            print("duplicate removed from {}".format(tempowner))
            if(x.owner != y.owner):
                x.owner = x.owner + " & {}".format(tempowner)
            
#for x in playlist:
#    x.printData()
    
shuffle(playlist)
shuffle(playlist)

iter = 0
# run the game
for x in playlist:
    command = input("(p)ause, (n)ext song, or (q)uit: ")
    if command == "p":
        system("./sp play")

    if command == "n":
        iter += 1
        print(iter)
        songuri = x.link
        print(songuri)
        system("./sp open {}".format(songuri))
        input("Press Enter to see who submitted the song.")
        print(x.owner)

    if command == "q":
        break

